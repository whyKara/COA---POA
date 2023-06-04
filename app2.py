import streamlit as st
from openpyxl import load_workbook
import sqlite3

# Connect to the database
conn = sqlite3.connect('db2.db')
cursor = conn.cursor()

# Function to insert data into the database
def insert_data_to_db(data):
    try:
        # Extract data from the input
        class_id = data['Class ID']
        test_id = data['Test ID']
        sub_id = data['Subject ID']
        question_id = data['Question ID']
        question_marks = data['Question Marks']
        co_attainment = data['CO Attainment']

        # Insert data into the DataMaster table
        cursor.execute('''
            INSERT INTO DataMaster (class_id, test_id, sub_id, question_id, question_marks, co_attainment)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (class_id, test_id, sub_id, question_id, question_marks, co_attainment))

        conn.commit()
        st.success("Data inserted successfully.")
    except Exception as e:
        st.error(f"Error occurred: {str(e)}")

# Read specific cells from the Excel file
def read_excel_file(file, sheet_name):
    try:
        # Load the Excel workbook
        workbook = load_workbook(file, read_only=True)

        # Select the sheet by name
        sheet = workbook[sheet_name]

        # Extract the required data from the sheet
        class_id = sheet['A8'].value
        subject_id = sheet['A10'].value

        # Check if the extracted values are None or empty strings
        if not class_id or not subject_id:
            st.warning("Class ID or Subject ID is not valid.")
            return

        data = {
            'Class ID': class_id,
            'Subject ID': subject_id
        }

        # Get the test IDs and their respective column indices
        test_ids = [sheet.cell(row=18, column=3).value, sheet.cell(row=18, column=12).value]
        test_columns = [3, 12]

        # Iterate over the tests and retrieve the questions, marks, and CO attainment values
        for test_id, col in zip(test_ids, test_columns):
            data['Test ID'] = test_id

            # Retrieve the questions, marks, and CO attainment values for the current test
            for row in range(19, 22):
                for col_offset in range(1, 5):
                    if col==3:
                        cell = sheet.cell(row=row, column=col + col_offset)
                    else:
                        cell=sheet.cell(row=row, column=col+ (col_offset-1))
                        
                    #question_id = cell.value
                    #question_marks = sheet.cell(row=22, column=col + col_offset).value
                    #co_attainment = sheet.cell(row=23, column=col + col_offset).value
                    if row==19:
                        question_marks=cell.value
                        data['Question Marks']=question_marks
                        #print(question_marks)
                    elif row==20:
                        co_attainment=cell.value
                        data['CO Attainment'] = co_attainment
                    elif row==21:
                        question_id=cell.value
                        data['Question ID'] = question_id
                    # Check if the question ID is not empty
                    #if question_id:
                        # Insert the data into the database
                        #data['Question ID'] = question_id
                        #data['Question Marks'] = question_marks
                        #data['CO Attainment'] = co_attainment

                        insert_data_to_db(data)
                    else:
                        # Break the loop if an empty cell is encountered
                        break

    except Exception as e:
        st.error(f"Error occurred: {str(e)}")

# Streamlit app
def main():
    st.title("Data Upload")
    st.write("---")

    # Page selection
    page = st.sidebar.selectbox("Select Page", ["Upload Excel File"])

    if page == "Upload Excel File":
        st.header("Upload Excel File")
        file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

        if file is not None:
            # Read the selected sheet from the uploaded Excel file
            workbook = load_workbook(file, read_only=True)
            sheet_names = workbook.sheetnames
            selected_sheet = st.selectbox("Select Sheet", sheet_names)
            read_excel_file(file, selected_sheet)

# Run the app
if __name__ == "__main__":
    main()
